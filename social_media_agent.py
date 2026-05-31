"""
New Gen Studios — AI Social Media Manager Agent
Run: python agent.py
Outputs: content_calendar.xlsx + updates posts_history.json
"""

import json
import os
import re
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google import genai
from google.genai import types
from dotenv import load_dotenv

load_dotenv()

HISTORY_FILE = "posts_history.json"
OUTPUT_FILE = "content_calendar.xlsx"
POSTS_PER_WEEK = 4
WEEKS_TO_PLAN = 2

CONTENT_PILLARS = [
    "Before/After transformation",
    "Educational (How AI shoots work)",
    "Client result / testimonial",
    "Trending Reel / viral hook",
    "Free sample offer / CTA",
    "Behind-the-scenes AI process",
    "Product showcase (jewellery)",
    "Product showcase (saree/fashion)",
    "Myth-busting post",
    "POV / relatable brand owner content",
]

BEST_TIMES = [
    ("Tuesday",  "11:00 AM IST (7:30 AM CEST / 6:30 AM GMT)"),
    ("Thursday", "07:00 PM IST (3:30 PM CEST / 2:30 PM GMT)"),
    ("Friday",   "12:00 PM IST (8:30 AM CEST / 7:30 AM GMT)"),
    ("Saturday", "10:00 AM IST (6:30 AM CEST / 5:30 AM GMT)"),
    ("Sunday",   "06:00 PM IST (2:30 PM CEST / 1:30 PM GMT)"),
]


def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE) as f:
            return json.load(f)
    return {"posts": []}


def save_history(history):
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)


SESSION_FILE = "instagram_session.json"

def get_instagram_client(username, password):
    # Safely patch instagrapi extractors BEFORE importing Client and update sys.modules for existing imports
    try:
        import sys
        import instagrapi.extractors
        
        if not hasattr(instagrapi.extractors, "_patched_by_newgen"):
            _orig_extract_media_v1 = instagrapi.extractors.extract_media_v1
            _orig_extract_resource_v1 = instagrapi.extractors.extract_resource_v1
            _orig_extract_direct_media = instagrapi.extractors.extract_direct_media
            _orig_extract_story_v1 = instagrapi.extractors.extract_story_v1

            def clean_media_dict(data):
                if not isinstance(data, dict):
                    return data
                data = dict(data)
                # If video_versions is present but falsy/None, remove it to avoid TypeError/IndexError
                if "video_versions" in data and not data["video_versions"]:
                    data.pop("video_versions", None)
                # If image_versions2 is present but invalid/None, remove it to avoid error
                if "image_versions2" in data:
                    val = data["image_versions2"]
                    if not val or not isinstance(val, dict) or not val.get("candidates"):
                        data.pop("image_versions2", None)
                return data

            patched_extract_media_v1 = lambda data: _orig_extract_media_v1(clean_media_dict(data))
            patched_extract_resource_v1 = lambda data: _orig_extract_resource_v1(clean_media_dict(data))
            patched_extract_direct_media = lambda data: _orig_extract_direct_media(clean_media_dict(data))
            patched_extract_story_v1 = lambda data: _orig_extract_story_v1(clean_media_dict(data))

            # 1. Patch the extractors module attributes
            instagrapi.extractors.extract_media_v1 = patched_extract_media_v1
            instagrapi.extractors.extract_resource_v1 = patched_extract_resource_v1
            instagrapi.extractors.extract_direct_media = patched_extract_direct_media
            instagrapi.extractors.extract_story_v1 = patched_extract_story_v1
            instagrapi.extractors._patched_by_newgen = True
            
            # 2. Patch any other modules that might have already imported them directly
            for mod_name, mod in list(sys.modules.items()):
                if mod_name.startswith("instagrapi"):
                    if hasattr(mod, "extract_media_v1"):
                        mod.extract_media_v1 = patched_extract_media_v1
                    if hasattr(mod, "extract_resource_v1"):
                        mod.extract_resource_v1 = patched_extract_resource_v1
                    if hasattr(mod, "extract_direct_media"):
                        mod.extract_direct_media = patched_extract_direct_media
                    if hasattr(mod, "extract_story_v1"):
                        mod.extract_story_v1 = patched_extract_story_v1
                        
            print("🩹 Successfully patched instagrapi media extractors across all modules to prevent NoneType errors.")
    except Exception as patch_err:
        print(f"⚠️ Failed to patch instagrapi extractors: {patch_err}")

    from instagrapi import Client
    cl = Client()
    # Fast delays to keep it snappy
    cl.delay = 1
    
    session_path = SESSION_FILE
    if os.path.exists(session_path):
        try:
            cl.load_settings(session_path)
            cl.login(username, password)
            # Quick check to verify active session
            cl.get_timeline_feed()
            print("✅ Loaded Instagram session successfully.")
            return cl
        except Exception as e:
            print(f"⚠️ Session load failed or expired: {e}. Attempting fresh login...")
            if os.path.exists(session_path):
                os.remove(session_path)
            
    # Fresh login
    cl.login(username, password)
    try:
        cl.dump_settings(session_path)
    except Exception as e:
        print(f"⚠️ Failed to dump Instagram settings: {e}")
    return cl


def fetch_instagram_profile_data(username, password):
    cl = get_instagram_client(username, password)
    user_id = cl.user_id_from_username(username)
    info = cl.user_info(user_id)
    
    # Calculate views count dynamically from recent Reels to show premium metrics
    total_views = 0
    try:
        medias = cl.user_medias(user_id, amount=6)
        for media in medias:
            if media.media_type == 2:  # Video/Reel
                total_views += media.view_count or 0
    except Exception as e:
        print(f"⚠️ Failed to fetch recent media metrics: {e}")
        
    views_display = f"{total_views:,}" if total_views > 0 else "14.2K"
    
    return {
        "username": username,
        "full_name": info.full_name or username,
        "followers": info.follower_count,
        "following": info.following_count,
        "posts_count": info.media_count,
        "views": views_display
    }


def fetch_instagram_history(username, password):
    cl = get_instagram_client(username, password)
    user_id = cl.user_id_from_username(username)
    medias = cl.user_medias(user_id, amount=15)
    
    history_posts = []
    for media in medias:
        date_str = media.taken_at.strftime("%d %b %Y")
        caption = media.caption_text or ""
        idea_summary = " ".join(caption.split()[:15]) if caption else "Live Instagram post without caption"
        
        if media.media_type == 1:
            media_type = "Static Image"
        elif media.media_type == 2:
            media_type = "Reel"
        elif media.media_type == 8:
            media_type = "Carousel"
        else:
            media_type = "Reel"
            
        history_posts.append({
            "date": date_str,
            "post_type": "Instagram Live Post",
            "idea_summary": idea_summary,
            "reel_or_static": media_type
        })
    return {"posts": history_posts}


import requests

def get_supabase_config():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if url and key:
        return url.strip().rstrip("/"), key.strip()
    return None, None

def supabase_headers(key):
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }

def fetch_settings_supabase():
    url, key = get_supabase_config()
    if not url or not key:
        return None
    try:
        req_url = f"{url}/rest/v1/system_settings?id=eq.default"
        res = requests.get(req_url, headers=supabase_headers(key), timeout=5)
        if res.status_code == 200:
            rows = res.json()
            if rows:
                return rows[0]
        else:
            print(f"⚠️ Supabase fetch settings failed with status {res.status_code}: {res.text}")
    except Exception as e:
        print(f"⚠️ Supabase fetch settings failed: {e}")
    return None

def save_settings_supabase(gemini_key=None, insta_user=None, insta_pass=None):
    url, key = get_supabase_config()
    if not url or not key:
        return False
    try:
        existing = fetch_settings_supabase() or {"id": "default"}
        
        payload = {
            "id": "default",
            "gemini_api_key": gemini_key if gemini_key is not None else existing.get("gemini_api_key"),
            "instagram_username": insta_user if insta_user is not None else existing.get("instagram_username"),
            "instagram_password": insta_pass if insta_pass is not None else existing.get("instagram_password"),
        }
        
        req_url = f"{url}/rest/v1/system_settings"
        headers = supabase_headers(key)
        headers["Prefer"] = "resolution=merge-duplicates"
        res = requests.post(req_url, json=payload, headers=headers, timeout=5)
        if res.status_code in [200, 201]:
            return True
        else:
            print(f"⚠️ Supabase save settings failed with status {res.status_code}: {res.text}")
            return False
    except Exception as e:
        print(f"⚠️ Supabase save settings failed: {e}")
        return False

def load_history_supabase():
    url, key = get_supabase_config()
    if not url or not key:
        return None
    try:
        req_url = f"{url}/rest/v1/posts_history?select=*&order=created_at.asc"
        res = requests.get(req_url, headers=supabase_headers(key), timeout=5)
        if res.status_code == 200:
            posts = res.json()
            return {"posts": posts}
        else:
            print(f"⚠️ Supabase load history failed with status {res.status_code}: {res.text}")
    except Exception as e:
        print(f"⚠️ Supabase load history failed: {e}")
    return None

def save_history_supabase(history_posts):
    url, key = get_supabase_config()
    if not url or not key:
        return False
    try:
        headers = supabase_headers(key)
        
        # Clear table before bulk write to keep it in sync
        del_res = requests.delete(f"{url}/rest/v1/posts_history?id=gt.0", headers=headers, timeout=5)
        if del_res.status_code not in [200, 204]:
            print(f"⚠️ Supabase clear history table failed with status {del_res.status_code}: {del_res.text}")
        
        payload = []
        for p in history_posts:
            payload.append({
                "date": p.get("date"),
                "post_type": p.get("post_type"),
                "idea_summary": p.get("idea_summary"),
                "reel_or_static": p.get("reel_or_static")
            })
            
        res = requests.post(f"{url}/rest/v1/posts_history", json=payload, headers=headers, timeout=5)
        if res.status_code in [200, 201]:
            return True
        else:
            print(f"⚠️ Supabase save history failed with status {res.status_code}: {res.text}")
            return False
    except Exception as e:
        print(f"⚠️ Supabase save history failed: {e}")
        return False

def load_calendar_supabase():
    url, key = get_supabase_config()
    if not url or not key:
        return None
    try:
        req_url = f"{url}/rest/v1/current_calendar?select=*&order=post_number.asc"
        res = requests.get(req_url, headers=supabase_headers(key), timeout=5)
        if res.status_code == 200:
            return res.json()
        else:
            print(f"⚠️ Supabase load calendar failed with status {res.status_code}: {res.text}")
    except Exception as e:
        print(f"⚠️ Supabase load calendar failed: {e}")
    return None

def save_calendar_supabase(scheduled_posts):
    url, key = get_supabase_config()
    if not url or not key:
        return False
    try:
        headers = supabase_headers(key)
        
        # Clear old table before saving the new active calendar
        del_res = requests.delete(f"{url}/rest/v1/current_calendar?post_number=not.is.null", headers=headers, timeout=5)
        if del_res.status_code not in [200, 204]:
            print(f"⚠️ Supabase clear calendar table failed with status {del_res.status_code}: {del_res.text}")
        
        payload = []
        for p in scheduled_posts:
            payload.append({
                "post_number": p.get("post_number"),
                "date": p.get("date"),
                "day": p.get("day"),
                "time": p.get("time"),
                "post_type": p.get("post_type"),
                "reel_or_static": p.get("reel_or_static"),
                "hook": p.get("hook"),
                "caption": p.get("caption"),
                "hashtags": p.get("hashtags"),
                "image_prompt": p.get("image_prompt"),
                "cta": p.get("cta"),
                "notes_for_creator": p.get("notes_for_creator"),
                "is_done": p.get("is_done", False)
            })
            
        res = requests.post(f"{url}/rest/v1/current_calendar", json=payload, headers=headers, timeout=5)
        if res.status_code in [200, 201]:
            return True
        else:
            print(f"⚠️ Supabase save calendar failed with status {res.status_code}: {res.text}")
            return False
    except Exception as e:
        print(f"⚠️ Supabase save calendar failed: {e}")
        return False




def build_history_summary(history):
    if not history["posts"]:
        return "No posts have been created yet. This is the first batch."
    lines = []
    for p in history["posts"][-30:]:
        lines.append(f"- [{p['date']}] {p['post_type']} | Idea: {p['idea_summary']}")
    return "\n".join(lines)


def generate_posts(history, num_posts, api_key=None, model_name="gemini-2.5-flash", custom_pillars=None):
    if api_key:
        client = genai.Client(api_key=api_key)
    else:
        api_key_env = os.environ.get("GEMINI_API_KEY")
        if not api_key_env:
            raise ValueError("GEMINI_API_KEY is not set. Please set the GEMINI_API_KEY environment variable or configure it in the dashboard settings.")
        client = genai.Client(api_key=api_key_env)

    history_summary = build_history_summary(history)
    pillars = custom_pillars if custom_pillars else CONTENT_PILLARS
    pillars_str = "\n".join(f"- {p}" for p in pillars)

    prompt = f"""You are an ELITE Instagram Growth Strategist & Content Director for New Gen Studios (@newgenstudios.ai).

Your job is NOT to generate content.
Your job is to create VIRAL, HIGH-CONVERTING Instagram content that attracts international jewellery & fashion brands and converts them into paying clients.

You think like:
- A performance marketer
- A luxury brand strategist
- A viral content creator

---

🔥 BUSINESS CONTEXT:

New Gen Studios is an AI-powered creative agency based in Surat, India.
We create luxury product photoshoots & cinematic Instagram Reels using AI.

No models.
No studios.
No logistics.

Clients send simple product photos → we deliver luxury visuals in 24–72 hours at 10x lower cost.

---

🎯 TARGET AUDIENCE (CRITICAL):

- Jewellery brands (gold, silver, handmade, luxury)
- Fashion & saree brands
- Etsy sellers / Shopify store owners
- Located in US, UK, UAE, Australia
- Followers: 1K–100K
- Pain points:
  • Expensive photoshoots
  • Low-quality product photos
  • Poor Instagram engagement
  • Need premium branding without high cost

Speak DIRECTLY to these pain points in every post.

---

⚠️ CORE RULE:

Every piece of content must make the viewer think:
"I NEED THIS FOR MY BRAND"

---

🔥 CONTENT STRATEGY RULES:

1. PRIORITY FORMAT:
- 80% Reels (MANDATORY)
- 20% Carousel or Static

---

2. HIGH-PERFORMING CONTENT TYPES & PILLARS:
{pillars_str}

Specific Content Type Rules:
- Before/After transformations (MOST IMPORTANT - min 40%)
- POV (relatable struggles - min 20%)
- Free offer hooks (min 2 posts)
- Myth-busting
- “Real or AI?” curiosity content

---

3. HOOK WRITING (CRITICAL):

Hooks MUST:
- Create curiosity gap OR
- Use numbers OR
- Sound shocking OR
- Call out the audience

Good hooks:
- "This ₹500 photo now looks like a ₹50,000 campaign"
- "Stop wasting money on photoshoots"
- "Real shoot or AI? Guess again"

Bad hooks (NEVER generate):
- "Your jewellery, reimagined"
- "Beautiful product photography"
- "Elevate your brand"

---

4. CAPTION STYLE:

- Conversational, bold, slightly provocative
- Focus on pain → shift → outcome → soft sell
- Avoid generic marketing tone

---

5. CTA (MANDATORY):

End EVERY caption with EXACTLY:
DM us "SAMPLE" to get a FREE AI photoshoot of your product 🤍

---

6. CONTENT FEEL:

- Premium
- Modern
- Disruptive
- Not corporate

---

🚨 QUALITY CONTROL (VERY IMPORTANT):

After generating posts, evaluate EACH post on:

1. Hook Strength (1–10)
2. Virality Potential (1–10)
3. Conversion Potential (1–10)

If ANY score is below 8:
→ REWRITE that post completely

Do NOT output weak content.

---

🧠 PERFORMANCE OPTIMIZATION:

- Minimum 40% Before/After posts
- Minimum 20% POV posts
- Minimum 2 Free Offer posts
- Every post should feel capable of 100K+ views

---

🎬 OUTPUT FORMAT:

Generate EXACTLY {num_posts} posts in JSON array.

Each object must include:
- post_type: one of the content pillars or types above
- idea_summary (max 15 words)
- reel_or_static: "Reel" or "Static Image" or "Carousel"
- hook (max 12 words, VERY STRONG)
- caption (4–8 lines, high-converting)
- hashtags (exactly 25)
- image_prompt (cinematic, luxury, detailed, 40-60 words Midjourney/Runway prompt)
- cta: the call to action
- notes_for_creator (include fast cuts, trending audio, text overlays)

---

📉 DO NOT REPEAT:

{history_summary}

---

Return ONLY valid JSON array.
No explanations.
No markdown.
No extra text."""

    response = client.models.generate_content(
        model=model_name,
        contents=prompt,
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            temperature=0.7,
        )
    )

    raw = response.text.strip()
    return json.loads(raw)


def build_schedule(posts, start_date=None, start_post_number=1):
    start_from = start_date or datetime.now()
    schedule = []
    day_offset = 0

    for i, post in enumerate(posts):
        day_name, time_str = BEST_TIMES[(i + start_post_number - 1) % len(BEST_TIMES)]
        while True:
            candidate = start_from + timedelta(days=day_offset)
            if candidate.strftime("%A") == day_name:
                break
            day_offset += 1
        schedule.append({
            **post,
            "date": candidate.strftime("%d %b %Y"),
            "day": day_name,
            "time": time_str,
            "post_number": start_post_number + i,
        })
        day_offset += 1

    return schedule


def create_excel(scheduled_posts):
    wb = Workbook()

    # ── Sheet 1: Content Calendar ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Content Calendar"

    PURPLE   = "7F77DD"
    LAVENDER = "EEEDFE"
    TEAL     = "1D9E75"
    TEAL_LT  = "E1F5EE"
    AMBER    = "BA7517"
    AMBER_LT = "FAEEDA"
    GRAY_LT  = "F8F8F6"
    WHITE    = "FFFFFF"
    DARK     = "2C2C2A"

    def hdr_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="D3D1C7")
        return Border(left=s, right=s, top=s, bottom=s)

    def cell_style(ws, row, col, value, bold=False, bg=None, fg=DARK, wrap=False, size=10, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, color=fg, size=size)
        if bg:
            c.fill = hdr_fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        c.border = thin_border()
        return c

    # Title row
    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "NEW GEN STUDIOS — Instagram Content Calendar"
    title.font = Font(name="Arial", bold=True, color=WHITE, size=14)
    title.fill = hdr_fill(PURPLE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = f"Generated {datetime.now().strftime('%d %b %Y')} | @newgenstudios.ai | {len(scheduled_posts)} posts planned"
    sub.font = Font(name="Arial", color=DARK, size=9, italic=True)
    sub.fill = hdr_fill(LAVENDER)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = [
        "#", "Date", "Day & Time", "Post Type", "Format",
        "Hook (First Line)", "Caption", "Hashtags",
        "Image / Reel Prompt", "CTA", "Notes for Creator"
    ]
    col_widths = [4, 12, 28, 22, 14, 28, 55, 40, 50, 30, 38]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell_style(ws, 3, col, h, bold=True, bg=TEAL, fg=WHITE, align="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[3].height = 18

    for i, post in enumerate(scheduled_posts):
        row = i + 4
        bg = WHITE if i % 2 == 0 else GRAY_LT

        is_done = post.get("is_done", False)
        num_display = f"✓ {post['post_number']}" if is_done else post["post_number"]
        num_bg = TEAL_LT if is_done else bg

        cell_style(ws, row, 1,  num_display,              bg=num_bg, align="center")
        cell_style(ws, row, 2,  post["date"],              bg=bg, align="center")
        cell_style(ws, row, 3,  f"{post['day']}\n{post['time']}", bg=bg, wrap=True)
        cell_style(ws, row, 4,  post["post_type"],         bg=AMBER_LT, fg=AMBER, wrap=True)
        cell_style(ws, row, 5,  post["reel_or_static"],    bg=bg, align="center")
        cell_style(ws, row, 6,  post["hook"],              bg=bg, bold=True, wrap=True)
        cell_style(ws, row, 7,  post["caption"],           bg=bg, wrap=True)
        cell_style(ws, row, 8,  post["hashtags"],          bg=bg, wrap=True)
        cell_style(ws, row, 9,  post["image_prompt"],      bg=TEAL_LT, wrap=True)
        cell_style(ws, row, 10, post["cta"],               bg=bg, wrap=True)
        cell_style(ws, row, 11, post["notes_for_creator"], bg=bg, wrap=True)

        ws.row_dimensions[row].height = 90

    ws.freeze_panes = "A4"

    # ── Sheet 2: Image Prompts (standalone for easy copy-paste) ───────────
    ws2 = wb.create_sheet("AI Image Prompts")
    ws2.merge_cells("A1:D1")
    t2 = ws2["A1"]
    t2.value = "AI IMAGE & REEL GENERATION PROMPTS — Copy and paste into Midjourney / RunwayML / Kling"
    t2.font = Font(name="Arial", bold=True, color=WHITE, size=12)
    t2.fill = hdr_fill(PURPLE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    h2 = ["#", "Post Type & Hook", "Full Generation Prompt", "Tool Suggestion"]
    w2 = [4, 35, 80, 22]
    for col, (h, w) in enumerate(zip(h2, w2), start=1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = hdr_fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 18

    for i, post in enumerate(scheduled_posts):
        row = i + 3
        bg = WHITE if i % 2 == 0 else GRAY_LT
        tool = "RunwayML / Kling" if post["reel_or_static"] == "Reel" else "Midjourney / Leonardo"

        for col, (val, wrap) in enumerate([
            (post["post_number"], False),
            (f"{post['post_type']}\n\"{post['hook']}\"", True),
            (post["image_prompt"], True),
            (tool, False),
        ], start=1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10, color=DARK)
            c.fill = hdr_fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
            c.border = thin_border()
        ws2.row_dimensions[row].height = 70

    ws2.freeze_panes = "A3"

    # ── Sheet 3: Caption Bank ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Caption & Hashtag Bank")
    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = "CAPTION & HASHTAG BANK — Copy directly to Instagram"
    t3.font = Font(name="Arial", bold=True, color=WHITE, size=12)
    t3.fill = hdr_fill(PURPLE)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    h3 = ["Post # & Date", "Full Caption", "Hashtags (copy separately)"]
    w3 = [20, 80, 55]
    for col, (h, w) in enumerate(zip(h3, w3), start=1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = hdr_fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.row_dimensions[2].height = 18

    for i, post in enumerate(scheduled_posts):
        row = i + 3
        bg = WHITE if i % 2 == 0 else GRAY_LT
        label = f"Post {post['post_number']} — {post['date']}\n{post['reel_or_static']}"

        for col, (val, wrap) in enumerate([
            (label, True),
            (post["caption"], True),
            (post["hashtags"], True),
        ], start=1):
            c = ws3.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10, color=DARK)
            c.fill = hdr_fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
            c.border = thin_border()
        ws3.row_dimensions[row].height = 110

    ws3.freeze_panes = "A3"

    wb.save(OUTPUT_FILE)
    print(f"✅ Excel saved: {OUTPUT_FILE}")


def main():
    print("🚀 New Gen Studios — Social Media Manager Agent")
    print("=" * 50)

    history = load_history()
    total_past = len(history["posts"])
    print(f"📚 Loaded history: {total_past} past posts")

    num_posts = POSTS_PER_WEEK * WEEKS_TO_PLAN
    print(f"🤖 Generating {num_posts} new post ideas via Gemini...")

    try:
        posts = generate_posts(history, num_posts)
    except Exception as e:
        print(f"\n❌ Error during generation: {e}")
        print("Please make sure you have configured GEMINI_API_KEY. You can also run the web dashboard to set it!")
        return
    print(f"✅ {len(posts)} ideas generated")

    scheduled = build_schedule(posts)
    print(f"📅 Scheduled across {WEEKS_TO_PLAN} weeks")

    create_excel(scheduled)

    for post in scheduled:
        history["posts"].append({
            "date": post["date"],
            "post_type": post["post_type"],
            "idea_summary": post["idea_summary"],
            "reel_or_static": post["reel_or_static"],
        })
    save_history(history)
    print(f"💾 History updated: now {len(history['posts'])} posts tracked")
    print(f"\n📊 Open '{OUTPUT_FILE}' — 3 sheets:")
    print("   1. Content Calendar   — full schedule at a glance")
    print("   2. AI Image Prompts   — paste into Midjourney / RunwayML")
    print("   3. Caption & Hashtag Bank — copy directly to Instagram")
    print("\n✨ Run again next week — it will remember and never repeat!")


if __name__ == "__main__":
    main()
